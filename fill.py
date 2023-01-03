import requests
import pandas as pd
import numpy as np
import random
from tqdm import tqdm
from bs4 import BeautifulSoup
from openpyxl import Workbook
import openpyxl

kept_columns = [" Prix de Vente TTC "]


def add_to_dict(D, origin):
    for x in D:
        origin[x] = origin.get(x,[]) + [D[x]]


def get_dewey_data(isbn):
    r = requests.get(f"http://classify.oclc.org/classify2/ClassifyDemo?search-standnum-txt={isbn}&startRec=0")
    soup = BeautifulSoup(r.content, 'html.parser')
    D = {}
    for title, value in zip(soup.find_all("dt"),soup.find_all("dd")):
        key, val = (title.string, value.string)
        if not key or not val:
            continue
        if key == "Author:":
            D["Auteur"] = " ".join(val.split(",")[::-1])
        elif key == "Title:":
            D["Titre"] = " ".join(val.split(",")[::-1])
    L = soup.find_all("td")
    if len(L) < 2:
        return D
    L = L[1]
    D["Dewey"] = L.string
    domains = []
    L = soup.find_all("table", attrs={"id":"subheadtbl"})
    if len(L) == 0:
        return D
    L = L[0]
    L = list(L.descendants)[16]
    for i in L:
        if str(type(i)) == "<class 'bs4.element.Tag'>":
            domains.append(list(list(i.descendants)[1].descendants)[0].string)
    D["Domaines"] = " / ".join(domains)
    return D


def get_book_data(isbn,verbose=False):
    r = requests.get(f"https://www.googleapis.com/books/v1/volumes?q={isbn}")
    ans = r.json()
    D = {}
    if ans["totalItems"] == 0:
        if verbose:
            print(f"Book with ISBN: {isbn} not found !")
    else:
        title, author, publisher, publishedDate, description = None,None,None,None,None
        if "title" in ans["items"][0]["volumeInfo"]:
            title = ans["items"][0]["volumeInfo"]["title"]
        if "subtitle" in ans["items"][0]["volumeInfo"]:
            title = title + ", " + ans["items"][0]["volumeInfo"]["subtitle"]
        if "authors" in ans["items"][0]["volumeInfo"]:
            author = ", ".join(ans["items"][0]["volumeInfo"]["authors"])
        if "publisher" in ans["items"][0]["volumeInfo"]:
            publisher = ans["items"][0]["volumeInfo"]["publisher"]
        if "publishedDate" in ans["items"][0]["volumeInfo"]:
            publishedDate = ans["items"][0]["volumeInfo"]["publishedDate"]
        if "description" in ans["items"][0]["volumeInfo"]:
            description = ans["items"][0]["volumeInfo"]["description"]
        D= {
            "ISBN":isbn,
            "Titre":title,
            "Auteur":author,
            "Editeur":publisher,
            "Date": publishedDate,
            "Description":description,
            "Dewey": None,
            "Domaines": None,
        }
    D2 = get_dewey_data(isbn)
    for x in D2:
        D[x] = D2[x]
    for x in D:
        D[x] = str(D[x]).strip() if D[x] is not None else ""
    return D


def autofill_df(df,verbose=False,kept_cols=None,correct=False,quick_mode=False,resume=False):
    if not resume:
        new_df = {}
    else:
        new_df = pd.read_csv("temp.csv").to_dict()
    for isbn, title, author, publisher, date in tqdm(list(zip(df["ISBN"], df["Titre"],df["Auteur"],df["Editeur"],df["Date"]))):
        if "ISBN" in new_df and isbn in new_df["ISBN"]:
            continue
        book_data = {"ISBN":isbn, "Titre":str(title), "Auteur":str(author),"Editeur":str(publisher), "Date":str(date)}
        if quick_mode and (str(isbn) != "nan" and str(title) != "nan" and str(author) != "nan" and str(publisher) != "nan" and str(date) != "nan"):
            d = book_data
        else:
            d = get_book_data(isbn)
        if not quick_mode:
            book_data["Dewey"] = d["Dewey"] if "Dewey" in d else ""
            book_data["Domaines"] = d["Domaines"] if "Domaines" in d else ""
        if not correct:
            book_data["Titre"] = "#"+d["Titre"] if book_data["Titre"] == "nan" and "Titre" in d else book_data["Titre"]
            book_data["Auteur"] = "#"+d["Auteur"] if book_data["Auteur"] == "nan" and "Auteur" in d else book_data["Auteur"]
            book_data["Editeur"] = "#"+d["Editeur"] if book_data["Editeur"] == "nan" and "Editeur" in d else book_data["Editeur"]
            book_data["Date"] = "#"+d["Date"] if book_data["Date"] == "nan" and "Date" in d else book_data["Date"]
        else:
            book_data["Titre"] = "%"+d["Titre"] if "Titre" in d else book_data["Titre"]
            book_data["Auteur"] = "%"+d["Auteur"] if "Auteur" in d else book_data["Auteur"]
            book_data["Editeur"] = "%"+d["Editeur"] if "Editeur" in d else book_data["Editeur"]
            book_data["Date"] = "%"+d["Date"] if "Date" in d else book_data["Date"]
        add_to_dict(book_data, new_df)
        new_df = pd.DataFrame(new_df)
        if kept_cols:
            new_df[kept_cols] = df[kept_cols]
        new_df.to_csv("temp.csv")
    return new_df


def output_completed_xlsx(input_path,output_path):
    d2 = pd.read_csv(input_path) 
    d3 = autofill_df(d2,kept_cols=kept_columns)
    wb = Workbook()
    ws = wb.active
    yellow = openpyxl.styles.colors.Color(rgb='F1F7B5')
    yellow_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=yellow)
    for i in range(d3.columns.shape[0]):
        c = ws.cell(row=1,column=i+1)
        c.value = d3.columns[i]
    for i in range(d3.values.shape[0]):
        for j in range(d3.values.shape[1]):
            c = ws.cell(row=i+2,column=j+1)
            c.value = str(d3.values[i,j])
            if c.value.startswith("#"):
                c.value = c.value[1:]
                c.fill = yellow_fill
    wb.save(output_path)
